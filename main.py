import csv
import configparser
from itertools import compress
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
import requests
from io import StringIO
from datetime import datetime
import pytz


def get_array_len(arr):
    len_arr = []
    for item in arr:
        len_arr.append(len(str(item)))
    return len_arr


def convert_time(str_time, str_fmt, str_timezone, out_fmt):
    if str_time != '':
        date_time_obj = datetime.strptime(str_time, str_fmt)
        date_time_obj = date_time_obj.astimezone(pytz.timezone(str_timezone))
        return date_time_obj.strftime(out_fmt)
    else:
        return ''


class TaigaTask:
    datetime_input_fmt = '%Y-%m-%d %H:%M:%S.%f%z'
    datetime_output_fmt = '%Y-%m-%d %H:%M:%S'
    datetime_timezone = 'Europe/Moscow'

    def __init__(self, ):
        self.obj_id = None
        self.obj_ref = None
        self.subject = None
        self.user_story = None
        self.assigned_to = None
        self.assigned_users = None
        self.status = None
        self.is_closed = None
        self.created_date = None
        self.modified_date = None
        self.finish_date = None
        self.due_date = None
        self.due_date_reason = None
        self.time_spent = None
        self.paid = None
        self.type = None

    def __str__(self):
        return str(self.__dict__)

    def __repr__(self):
        return str(self.__dict__)

    def get_vars(self):
        return list(self.__dict__.values())

    def get_vars_len(self):
        return get_array_len(self.get_vars())

    def get_var_names(self):
        return list(self.__dict__.keys())

    def get_var_names_len(self):
        return get_array_len(self.get_var_names())

    def convert_times(self):
        self.created_date = convert_time(self.created_date,
                                         self.datetime_input_fmt, self.datetime_timezone, self.datetime_output_fmt)
        self.modified_date = convert_time(self.modified_date,
                                          self.datetime_input_fmt, self.datetime_timezone, self.datetime_output_fmt)
        self.finish_date = convert_time(self.finish_date,
                                        self.datetime_input_fmt, self.datetime_timezone, self.datetime_output_fmt)
        self.due_date = convert_time(self.due_date,
                                     self.datetime_input_fmt, self.datetime_timezone, self.datetime_output_fmt)

    def set_user_story(self, data_arr):
        self.obj_id, \
        self.obj_ref, \
        self.subject, \
        self.assigned_to, \
        self.assigned_users, \
        self.status, \
        self.is_closed, \
        self.created_date, \
        self.modified_date, \
        self.finish_date, \
        self.due_date, \
        self.due_date_reason, \
        self.time_spent, \
        self.paid \
            = data_arr
        self.type = 'US'
        self.created_date = self.created_date if True else False
        self.convert_times()
        return self

    def set_task(self, data_arr):
        self.obj_id, \
        self.obj_ref, \
        self.subject, \
        self.user_story, \
        self.assigned_to, \
        self.status, \
        self.is_closed, \
        self.created_date, \
        self.modified_date, \
        self.finish_date, \
        self.due_date, \
        self.due_date_reason, \
        self.time_spent, \
        self.paid \
            = data_arr
        self.type = 'TK'
        self.convert_times()
        return self

    def set_issue(self, data_arr):
        self.obj_id, \
        self.obj_ref, \
        self.subject, \
        self.assigned_to, \
        self.status, \
        self.is_closed, \
        self.created_date, \
        self.modified_date, \
        self.finish_date, \
        self.due_date, \
        self.due_date_reason, \
        self.time_spent, \
        self.paid \
            = data_arr
        self.type = 'IS'
        self.convert_times()
        return self


def read_csv_from_file(file):
    with open(file, newline='', encoding='UTF-8') as csvfile:
        reader_obj = csv.reader(csvfile, delimiter=',')
        return process_tasks_from_rows(reader_obj)


def read_csv_from_link(link):
    csvfile = StringIO(link)
    reader_obj = csv.reader(csvfile, delimiter=',')
    return process_tasks_from_rows(reader_obj)


def process_tasks_from_rows(rows):
    filetype = None
    first_row = True
    col_filter = []
    out_arr = []
    for row in rows:
        if first_row:
            col_names_arr = row
            col_filter = [True if item in fields_list else False for item in col_names_arr]
            if 'assigned_users' in col_names_arr:
                filetype = 'user_stories'
            elif 'user_story' in col_names_arr:
                filetype = 'tasks'
            else:
                filetype = 'issues'

            first_row = False

        else:
            filtered_row = list(compress(row, col_filter))
            if filetype == 'user_stories':
                row_obj = TaigaTask().set_user_story(filtered_row)
            elif filetype == 'tasks':
                row_obj = TaigaTask().set_task(filtered_row)
            else:
                row_obj = TaigaTask().set_issue(filtered_row)

            out_arr.append(row_obj)

    return out_arr


def export_to_excel(taiga_arr, out_file):
    wb = Workbook()
    ws = wb.active
    if not taiga_arr:
        print('Nothing to save!')
        return False
    ws.append(taiga_arr[0].get_var_names())
    max_len = taiga_arr[0].get_var_names_len()

    for taiga_entry in taiga_arr:
        curr_row_len = taiga_entry.get_vars_len()

        for i in range(len(max_len)):
            if max_len[i] < curr_row_len[i]:
                max_len[i] = curr_row_len[i]

        ws.append(taiga_entry.get_vars())

    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, index=get_column_letter(col),
                                                             width=max_len[col - 1] + 2)

    ws.column_dimensions = dim_holder
    wb.save(out_file)


if __name__ == '__main__':
    config = configparser.ConfigParser()
    config.read('taiga.ini')

    # Global
    fields_list = set(config['global']['fields'].split(','))
    read_from_files = config['global'].getboolean('read_from_files')
    read_from_links = config['global'].getboolean('read_from_links')

    datetime_input_fmt = config['global']['datetime_input_fmt']
    datetime_output_fmt = config['global']['datetime_output_fmt']
    datetime_timezone = config['global']['datetime_timezone']

    curr_datetime = datetime.now()
    output_filename = config['global']['output_filename']
    output_filename = curr_datetime.strftime(output_filename)

    # Filters
    filter_user = config['filters']['filter_user']
    not_paid_only = config['filters'].getboolean('not_paid_only')
    with_filled_time_only = config['filters'].getboolean('with_filled_time_only')
    closed_only = config['filters'].getboolean('closed_only')
    closed_statuses_list = set(config['filters']['closed_statuses'].split(','))

    # Files & links
    files = set(config['files'].values())
    links = set(config['links'].values())

    TaigaTask.datetime_input_fmt = datetime_input_fmt
    TaigaTask.datetime_output_fmt = datetime_output_fmt
    TaigaTask.datetime_timezone = datetime_timezone

    out = []
    if read_from_files:
        for taiga_file in files:
            out = [*out, *read_csv_from_file(taiga_file)]

    if read_from_links:
        for taiga_file_link in links:
            resp = requests.get(taiga_file_link)
            out = [*out, *read_csv_from_link(resp.text)]

    if filter_user:
        out = [x for x in out if x.assigned_to == filter_user]
        output_filename = output_filename + '_' + filter_user

    if not_paid_only:
        out = [x for x in out if x.paid != 'True']
        output_filename = output_filename + '_not-paid'

    if with_filled_time_only:
        out = [x for x in out if x.time_spent != '']
        output_filename = output_filename + '_w-times'

    if closed_only:
        out = [x for x in out if x.status in closed_statuses_list]
        output_filename = output_filename + '_closed'

    # print(out)
    export_to_excel(out, output_filename + '.xlsx')
